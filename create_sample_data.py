#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
샘플 데이터 생성 스크립트
data/ 폴더에 3개의 샘플 엑셀 파일을 생성합니다.
  - s1_attendance.xlsx  : S1 출퇴근 기록 샘플
    - ecount_employees.xlsx : ECOUNT 사원정보 샘플
      - hourly_rates.xlsx   : 시급 기준표 샘플

      실행: python create_sample_data.py
      """

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs("data", exist_ok=True)

# ─────────────────────────────────────────
# 공통 스타일 헬퍼
# ─────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="AAAAAA")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, col_widths):
      """헤더 행 스타일 + 열 너비 설정"""
      for cell in ws[1]:
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = CENTER
                cell.border    = BORDER
            for row in ws.iter_rows(min_row=2):
                      for cell in row:
                                    cell.font      = BODY_FONT
                                    cell.alignment = CENTER
                                    cell.border    = BORDER
                            for i, w in enumerate(col_widths, 1):
                                      ws.column_dimensions[get_column_letter(i)].width = w
                                  ws.row_dimensions[1].height = 22

# ─────────────────────────────────────────
# 1. s1_attendance.xlsx
#    컬럼: 사번 | 이름 | 일자 | 시간
#    S1 단말기는 출근·퇴근을 각각 별도 행으로 기록
#    → 하루 2행(출근시간, 퇴근시간)
# ─────────────────────────────────────────
attendance_rows = [
      # 사번    이름      일자          시간
    ("EMP001","홍길동","2026-04-01","09:00"),
      ("EMP001","홍길동","2026-04-01","18:10"),
      ("EMP001","홍길동","2026-04-02","08:55"),
      ("EMP001","홍길동","2026-04-02","18:30"),
      ("EMP001","홍길동","2026-04-03","09:00"),
      ("EMP001","홍길동","2026-04-03","20:05"),  # 연장
      ("EMP001","홍길동","2026-04-06","09:00"),
      ("EMP001","홍길동","2026-04-06","18:00"),
      ("EMP001","홍길동","2026-04-07","09:00"),
      ("EMP001","홍길동","2026-04-07","18:00"),
      ("EMP001","홍길동","2026-04-08","09:00"),
      ("EMP001","홍길동","2026-04-08","18:00"),
      ("EMP001","홍길동","2026-04-09","09:00"),
      ("EMP001","홍길동","2026-04-09","18:00"),
      ("EMP001","홍길동","2026-04-10","09:00"),
      ("EMP001","홍길동","2026-04-10","18:00"),
      ("EMP001","홍길동","2026-04-13","09:00"),
      ("EMP001","홍길동","2026-04-13","22:30"),  # 야간
      ("EMP001","홍길동","2026-04-14","09:00"),
      ("EMP001","홍길동","2026-04-14","18:00"),
      ("EMP002","김영희","2026-04-01","08:00"),
      ("EMP002","김영희","2026-04-01","17:00"),
      ("EMP002","김영희","2026-04-02","08:05"),
      ("EMP002","김영희","2026-04-02","17:00"),
      ("EMP002","김영희","2026-04-03","08:00"),
      ("EMP002","김영희","2026-04-03","17:00"),
      ("EMP002","김영희","2026-04-06","08:00"),
      ("EMP002","김영희","2026-04-06","17:00"),
      ("EMP002","김영희","2026-04-07","08:00"),
      ("EMP002","김영희","2026-04-07","17:00"),
      ("EMP002","김영희","2026-04-08","08:00"),
      ("EMP002","김영희","2026-04-08","17:00"),
      ("EMP002","김영희","2026-04-09","08:00"),
      ("EMP002","김영희","2026-04-09","17:00"),
      ("EMP002","김영희","2026-04-10","08:00"),
      ("EMP002","김영희","2026-04-10","17:00"),
      ("EMP002","김영희","2026-04-13","08:00"),
      ("EMP002","김영희","2026-04-13","17:00"),
      ("EMP002","김영희","2026-04-14","08:00"),
      ("EMP002","김영희","2026-04-14","17:00"),
      ("EMP003","이철수","2026-04-01","22:00"),
      ("EMP003","이철수","2026-04-02","06:00"),  # 야간근무 (날짜 넘김)
      ("EMP003","이철수","2026-04-02","22:00"),
      ("EMP003","이철수","2026-04-03","06:00"),
      ("EMP003","이철수","2026-04-03","22:00"),
      ("EMP003","이철수","2026-04-04","06:00"),
      ("EMP003","이철수","2026-04-06","22:00"),
      ("EMP003","이철수","2026-04-07","06:00"),
      ("EMP003","이철수","2026-04-07","22:00"),
      ("EMP003","이철수","2026-04-08","06:00"),
      ("EMP003","이철수","2026-04-08","22:00"),
      ("EMP003","이철수","2026-04-09","06:00"),
      ("EMP003","이철수","2026-04-09","22:00"),
      ("EMP003","이철수","2026-04-10","06:00"),
      ("EMP003","이철수","2026-04-13","22:00"),
      ("EMP003","이철수","2026-04-14","06:00"),
]

wb1 = Workbook()
ws1 = wb1.active
ws1.title = "출퇴근기록"
ws1.append(["사번", "이름", "일자", "시간"])
for row in attendance_rows:
      ws1.append(list(row))
style_sheet(ws1, [10, 10, 14, 10])
wb1.save("data/s1_attendance.xlsx")
print("[완료] data/s1_attendance.xlsx 생성")

# ─────────────────────────────────────────
# 2. ecount_employees.xlsx
#    컬럼: 사번 | 이름 | 부서 | 직책 | 입사일 | 퇴사일
# ─────────────────────────────────────────
employee_rows = [
      ("EMP001","홍길동","생산부","사원","2024-03-01",""),
      ("EMP002","김영희","물류부","주임","2023-06-15",""),
      ("EMP003","이철수","경비팀","사원","2025-01-10",""),
      ("EMP004","박민준","생산부","사원","2025-09-01","2026-02-28"),  # 퇴사자
      ("EMP005","최수진","물류부","사원","2024-11-01",""),
]

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "사원정보"
ws2.append(["사번", "이름", "부서", "직책", "입사일", "퇴사일"])
for row in employee_rows:
      ws2.append(list(row))
style_sheet(ws2, [10, 10, 12, 10, 14, 14])
wb2.save("data/ecount_employees.xlsx")
print("[완료] data/ecount_employees.xlsx 생성")

# ─────────────────────────────────────────
# 3. hourly_rates.xlsx
#    컬럼: 사번 | 이름 | 시급
# ─────────────────────────────────────────
rate_rows = [
      ("EMP001","홍길동", 10030),
      ("EMP002","김영희", 11000),
      ("EMP003","이철수", 10500),
      ("EMP005","최수진", 10030),
]

wb3 = Workbook()
ws3 = wb3.active
ws3.title = "시급기준표"
ws3.append(["사번", "이름", "시급"])
for row in rate_rows:
      ws3.append(list(row))
style_sheet(ws3, [10, 10, 12])

# 시급 셀 숫자 서식
for r in range(2, ws3.max_row + 1):
      ws3.cell(row=r, column=3).number_format = "#,##0"

wb3.save("data/hourly_rates.xlsx")
print("[완료] data/hourly_rates.xlsx 생성")

print()
print("=== 샘플 데이터 생성 완료 ===")
print("data/ 폴더에서 파일을 확인하세요.")
print("이후 python main.py 를 실행하면 급여 계산이 진행됩니다.")
